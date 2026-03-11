"""
Multi-model comparison charts: error-bar plots and radar grids.

All functions accept a ``models`` dict ``{display_name: results_dir}``
so the model list is **never hardcoded**. Colors are auto-assigned from
a 20-colour palette (Macaron tones).

Usage (programmatic)::

    from empa.visualization.comparison import discover_models, plot_errorbar_bars, plot_radar_grid
    models = discover_models("results/benchmark_runs/epm-bench/")
    plot_errorbar_bars(models, output_dir="output/")
    plot_radar_grid(models, output_dir="output/")

Usage (CLI)::

    empa compare results/benchmark_runs/epm-bench/ --chart all
    empa compare results/benchmark_runs/epm-bench/ --models "gpt-4o,claude-3.5-sonnet"
"""

from __future__ import annotations

from math import ceil, pi
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

from empa.evaluation.epmq import (
    W_EFFICIENCY,
    W_OUTCOME,
    W_STABILITY,
    calculate_epmq_indices,
)

# ---------------------------------------------------------------------------
# Colour palette (20 Macaron tones — automatically cycled)
# ---------------------------------------------------------------------------
_PALETTE = [
    "#E8788A", "#E88FAB", "#E29A8E", "#DBA95E", "#D6CF6A",
    "#9AAAE0", "#A3D48C", "#80C9A8", "#7ACDE8", "#F0B490",
    "#B89EE8", "#C0B0D8", "#E89B9B", "#A8B8E0", "#C07A38",
    "#4A9050", "#3088B8", "#C82848", "#5A6EB8", "#6A4E98",
]

IDX_COLUMNS = [
    "Idx_RDI", "Idx_Etot", "Idx_Snet",
    "Idx_Rho", "Idx_Sproj", "Idx_Tau",
    "Idx_Rpos", "Idx_Align", "Idx_Pen",
]

RAW_METRICS_MAPPING = {
    "Distance_Improvement%": r"RDI (%)",
    "Total_Energy": r"$E_{\mathrm{total}}$",
    "Total_Net_Score": r"$S_{\mathrm{net}}$",
    "Net_Score_Per_Turn": r"$\rho$",
    "Avg_Effective_Projection": r"$S_{\mathrm{proj}}$",
    "Tortuosity": r"$\tau$",
    "Positive_Energy_Ratio%": r"$R_{\mathrm{pos}}$ (%)",
    "Avg_Alignment": r"$\overline{\cos\theta}$",
    "Penalty_Rate": r"$R_{\mathrm{pen}}$",
}

METRIC_GROUPS: List[Tuple[str, List[str]]] = [
    ("Outcome Quality", ["Distance_Improvement%", "Total_Energy", "Total_Net_Score"]),
    ("Process Efficiency", ["Net_Score_Per_Turn", "Avg_Effective_Projection", "Tortuosity"]),
    ("Process Stability", ["Positive_Energy_Ratio%", "Avg_Alignment", "Penalty_Rate"]),
]

# Column name mapping for backward compatibility with old Chinese-header CSVs
_CN_TO_EN = {
    "案例ID": "Case_ID", "主导轴": "Dominant_Axis", "难度": "Difficulty", "类别": "Category",
    "SP共情阈值": "SP_Empathy_Threshold", "SP情感需求等级": "SP_Emotional_Priority",
    "SP动机需求等级": "SP_Motivational_Priority", "SP认知需求等级": "SP_Cognitive_Priority",
    "C赤字": "C_Deficit", "A赤字": "A_Deficit", "P赤字": "P_Deficit", "总赤字": "Total_Deficit",
    "成功": "Success", "轮次": "Turns", "终止类型": "Termination_Type",
    "初始距离": "Initial_Distance", "最终距离": "Final_Distance",
    "距离改善率%": "Distance_Improvement%", "累积能量": "Total_Energy",
    "能量达标率%": "Energy_Achievement%", "能量盈余": "Energy_Surplus",
    "几何胜利": "Geometric_Win", "位置胜利": "Positional_Win", "能量胜利": "Energetic_Win",
    "C净分": "C_Net_Score", "A净分": "A_Net_Score", "P净分": "P_Net_Score",
    "总净分": "Total_Net_Score", "总Prog_sum": "Total_Prog_sum", "总Neg_sum": "Total_Neg_sum",
    "净分/轮": "Net_Score_Per_Turn", "Prog/轮": "Prog_Per_Turn", "Neg/轮": "Neg_Per_Turn",
    "路径曲率": "Tortuosity", "有效投影": "Avg_Effective_Projection",
    "正能量占比%": "Positive_Energy_Ratio%", "平均对齐度": "Avg_Alignment",
    "惩罚率": "Penalty_Rate", "能量波动方差": "Delta_E_Variance",
    "连续负能量最长序列": "Max_Negative_Streak", "失败类型": "Failure_Type",
    "平均值": "Mean", "标准差": "Std", "最小值": "Min", "最大值": "Max",
}

# Value translations
_VAL_DIFF = {"较易": "Easy", "中等": "Medium", "困难": "Hard", "极难": "Very Hard"}
_VAL_CAT = {
    "休闲娱乐": "Leisure", "身心健康": "Health", "生活状况": "Life",
    "观念认同": "Values", "人际关系": "Relations", "生涯发展": "Career",
}
_VAL_SP = {"高": "High", "中": "Medium", "低": "Low"}


def _normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize a descriptive_statistics DataFrame to English columns/values."""
    df = df.rename(columns=_CN_TO_EN)
    if "Difficulty" in df.columns:
        df["Difficulty"] = df["Difficulty"].map(lambda v: _VAL_DIFF.get(v, v))
    if "Category" in df.columns:
        df["Category"] = df["Category"].map(lambda v: _VAL_CAT.get(v, v))
    for sp_col in ("SP_Empathy_Threshold", "SP_Emotional_Priority",
                    "SP_Motivational_Priority", "SP_Cognitive_Priority"):
        if sp_col in df.columns:
            df[sp_col] = df[sp_col].map(lambda v: _VAL_SP.get(v, v))
    if "Success" in df.columns:
        df["Success"] = df["Success"].map(lambda v: {"✅": "Yes", "❌": "No"}.get(v, v))
    return df


# ---------------------------------------------------------------------------
# Model discovery
# ---------------------------------------------------------------------------

def discover_models(
    base_dir: str | Path,
    *,
    filter_names: Optional[Sequence[str]] = None,
) -> Dict[str, Path]:
    """Auto-discover model result directories under *base_dir*.

    A directory is recognised as a model if it contains
    ``descriptive_statistics.csv``.

    Parameters
    ----------
    base_dir : path-like
        Parent directory containing one sub-folder per model.
    filter_names : sequence of str, optional
        If given, only include directories whose name contains one of these
        substrings (case-insensitive).

    Returns
    -------
    dict
        ``{display_name: Path}`` ordered by EPM-Index descending.
    """
    base = Path(base_dir)
    candidates: Dict[str, Path] = {}

    for d in sorted(base.iterdir()):
        if not d.is_dir():
            continue
        csv = d / "descriptive_statistics.csv"
        if not csv.exists():
            continue
        if filter_names:
            if not any(fn.lower() in d.name.lower() for fn in filter_names):
                continue
        display = _dir_to_display(d.name)
        candidates[display] = d

    # Sort by EPM-Index descending
    scored: List[Tuple[str, Path, float]] = []
    for name, path in candidates.items():
        try:
            idx = _quick_epm_index(path / "descriptive_statistics.csv")
        except Exception:
            idx = 0.0
        scored.append((name, path, idx))
    scored.sort(key=lambda t: t[2], reverse=True)

    return {name: path for name, path, _ in scored}


def _dir_to_display(dirname: str) -> str:
    """Heuristic: turn a folder name into a human-readable model name."""
    name = dirname.split("_resampled")[0]
    parts = name.split("_")
    cleaned: list[str] = []
    for p in parts:
        if p.isdigit() and len(p) >= 6:
            continue
        cleaned.append(p)
    raw = "-".join(cleaned)

    _KNOWN: dict[str, str] = {
        "gpt-4o": "GPT-4o",
        "gpt-5.2-pro": "GPT-5.2 Pro",
        "claude-3.5-sonnet": "Claude 3.5 Sonnet",
        "claude-4.6-opus": "Claude 4.6 Opus",
        "gemini-2.5-pro": "Gemini 2.5 Pro",
        "gemini-3-pro-preview": "Gemini 3 Pro Preview",
        "deepseek-chat-v3-0324": "DeepSeek V3",
        "seed-1.6": "Seed 1.6",
        "seed-2.0": "Seed 2.0",
        "kimi-k2-0905": "Kimi k2",
        "doubao-1.5-character": "Doubao 1.5",
        "qwen3-235b-a22b-2507": "Qwen 3 235B",
        "qwen3-32b": "Qwen 3 32B",
        "qwen3-8b": "Qwen 3 8B",
        "llama-3.1-8b-instruct": "Llama 3.1 8B",
        "llama-3.3-70b-instruct": "Llama 3.3 70B",
    }
    if raw.lower() in _KNOWN:
        return _KNOWN[raw.lower()]

    return raw.replace("-", " ").title()


def _quick_epm_index(csv_path: Path) -> float:
    df = _normalize_df(pd.read_csv(csv_path))
    df = df[~df["Case_ID"].isin(["Mean", "Std", "Min", "Max"])]
    if df.empty:
        return 0.0
    means = {col: df[col].astype(float).mean() for col in IDX_COLUMNS if col in df.columns}
    outcome = (means.get("Idx_RDI", 0) + means.get("Idx_Etot", 0) + means.get("Idx_Snet", 0)) / 3
    efficiency = (means.get("Idx_Rho", 0) + means.get("Idx_Sproj", 0) + means.get("Idx_Tau", 0)) / 3
    stability = (means.get("Idx_Rpos", 0) + means.get("Idx_Align", 0) + means.get("Idx_Pen", 0)) / 3
    return 0.4 * outcome + 0.2 * efficiency + 0.4 * stability


def _assign_colors(model_names: Sequence[str]) -> Dict[str, str]:
    return {name: _PALETTE[i % len(_PALETTE)] for i, name in enumerate(model_names)}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_case_df(csv_path: Path, model_name: str) -> pd.DataFrame:
    df = _normalize_df(pd.read_csv(csv_path))
    df = df[~df["Case_ID"].isin(["Mean", "Std", "Min", "Max"])].copy()
    df["Model"] = model_name
    return df


def _load_all(models: Dict[str, Path]) -> pd.DataFrame:
    frames = []
    for name, path in models.items():
        csv = path / "descriptive_statistics.csv"
        if csv.exists():
            frames.append(_load_case_df(csv, name))
    if not frames:
        raise FileNotFoundError("No descriptive_statistics.csv found in any model directory")
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Error-bar bar charts (3x3 grid)
# ---------------------------------------------------------------------------

def plot_errorbar_bars(
    models: Dict[str, Path],
    output_dir: str | Path = ".",
    *,
    theme_color: str = "#D82E4D",
) -> Path:
    """Generate a 3x3 Mean +/- SE error-bar chart comparing models.

    Returns the path of the saved PNG.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    _apply_academic_style(plt)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    long_df = _load_all(models)
    model_names = list(models.keys())
    colors = _assign_colors(model_names)

    fig, axes = plt.subplots(3, 3, figsize=(16, 13))

    for row_idx, (group_name, metrics) in enumerate(METRIC_GROUPS):
        for col_idx, metric_col in enumerate(metrics):
            ax = axes[row_idx, col_idx]
            if metric_col not in long_df.columns:
                ax.text(0.5, 0.5, f"Missing: {metric_col}", ha="center", va="center")
                continue

            means, stds = [], []
            for m in model_names:
                vals = long_df.loc[long_df["Model"] == m, metric_col].astype(float)
                means.append(vals.mean())
                stds.append(vals.std())

            x_pos = np.arange(len(model_names))
            bars = ax.bar(x_pos, means, alpha=0.6, width=0.7, zorder=1)

            np.random.seed(42)
            for i, m in enumerate(model_names):
                mv = long_df.loc[long_df["Model"] == m, metric_col].astype(float).values
                jitter = np.random.uniform(-0.15, 0.15, size=len(mv))
                ax.scatter(x_pos[i] + jitter, mv, color=colors[m],
                           edgecolor="white", linewidth=0.5, s=15, alpha=0.8, zorder=10)

            ax.errorbar(x_pos, means, yerr=stds, fmt="none",
                        ecolor="#555", elinewidth=1, capsize=3, alpha=0.8, zorder=20)

            for i, bar in enumerate(bars):
                bar.set_color(colors[model_names[i]])
                bar.set_edgecolor("#555")
                bar.set_linewidth(0.8)

            ax.set_title(RAW_METRICS_MAPPING.get(metric_col, metric_col),
                         fontsize=12, fontweight="bold", pad=8, color=theme_color)
            short = [_short_name(m) for m in model_names]
            ax.set_xticks(x_pos)
            ax.set_xticklabels(short, fontsize=max(6, 9 - len(model_names) // 5),
                               color=theme_color, fontweight="bold", rotation=30, ha="right")
            ax.tick_params(axis="y", colors=theme_color, labelsize=9)
            for sp in ax.spines.values():
                sp.set_edgecolor(theme_color)
            ymin, ymax = ax.get_ylim()
            if ymin > 0:
                ax.set_ylim(bottom=0)
            ax.axhline(0, color=theme_color, lw=0.8, ls="--", zorder=2, alpha=0.5)
            ax.grid(axis="y", ls="--", alpha=0.3, color=theme_color)
            if col_idx == 0:
                ax.set_ylabel(group_name, fontsize=14, fontweight="bold",
                              labelpad=12, color=theme_color)

    fig.suptitle("EPM-Q Core Metrics (Mean ± Std)", fontsize=18,
                 fontweight="bold", y=0.99, color=theme_color)

    handles = [mpatches.Patch(facecolor=colors[m], edgecolor="black", lw=0.8,
                              label=_short_name(m), alpha=0.85) for m in model_names]
    ncol = min(7, len(model_names))
    fig.legend(handles=handles, loc="upper center", ncol=ncol,
               bbox_to_anchor=(0.5, 0.955), frameon=False, fontsize=10,
               labelcolor=theme_color)
    fig.tight_layout(rect=[0, 0, 1, 0.90])

    out = output_dir / "fig_epmq_errorbar_bars.png"
    plt.savefig(str(out), dpi=300, bbox_inches="tight")
    plt.savefig(str(out).replace(".png", ".pdf"), bbox_inches="tight")
    plt.close(fig)
    print(f"✅ Error-bar charts: {out}")
    return out


# ---------------------------------------------------------------------------
# Radar grid (small multiples)
# ---------------------------------------------------------------------------

_RADAR_CONFIGS = [
    {
        "id": "categories",
        "filename": "fig_epmq_radar_grid_categories.png",
        "title": "Scenario Categories Profiles",
        "subtitle": (
            "Each axis represents a scenario category: Leisure, Health, Life, Values, Relations, Career."
        ),
        "labels": ["Leisure", "Health", "Life", "Values", "Relations", "Career"],
        "filters": [
            lambda df: df[df["Category"] == "Leisure"],
            lambda df: df[df["Category"] == "Health"],
            lambda df: df[df["Category"] == "Life"],
            lambda df: df[df["Category"] == "Values"],
            lambda df: df[df["Category"] == "Relations"],
            lambda df: df[df["Category"] == "Career"],
        ],
    },
    {
        "id": "mechanism",
        "filename": "fig_epmq_radar_grid_mechanism.png",
        "title": "Mechanism Stress Test Profiles",
        "subtitle": (
            "Axes = empathy mechanism (A/C/P) × difficulty (Routine vs Hard)."
        ),
        "labels": ["A-Routine", "A-Hard", "C-Routine", "C-Hard", "P-Routine", "P-Hard"],
        "filters": [
            lambda df: df[(df["Dominant_Axis"] == "A") & (df["Difficulty"].isin(["Easy", "Medium"]))],
            lambda df: df[(df["Dominant_Axis"] == "A") & (df["Difficulty"].isin(["Hard", "Very Hard"]))],
            lambda df: df[(df["Dominant_Axis"] == "C") & (df["Difficulty"].isin(["Easy", "Medium"]))],
            lambda df: df[(df["Dominant_Axis"] == "C") & (df["Difficulty"].isin(["Hard", "Very Hard"]))],
            lambda df: df[(df["Dominant_Axis"] == "P") & (df["Difficulty"].isin(["Easy", "Medium"]))],
            lambda df: df[(df["Dominant_Axis"] == "P") & (df["Difficulty"].isin(["Hard", "Very Hard"]))],
        ],
    },
    {
        "id": "persona",
        "filename": "fig_epmq_radar_grid_persona.png",
        "title": "Persona Resilience Profiles",
        "subtitle": (
            "Axes = dominant need (A/C/P) × empathy threshold (Rec=low-mid, Def=high)."
        ),
        "labels": ["A-Rec", "A-Def", "C-Rec", "C-Def", "P-Rec", "P-Def"],
        "filters": [
            lambda df: df[(df["SP_Emotional_Priority"] == "High") & (df["SP_Empathy_Threshold"].isin(["Low", "Medium"]))],
            lambda df: df[(df["SP_Emotional_Priority"] == "High") & (df["SP_Empathy_Threshold"] == "High")],
            lambda df: df[(df["SP_Cognitive_Priority"] == "High") & (df["SP_Empathy_Threshold"].isin(["Low", "Medium"]))],
            lambda df: df[(df["SP_Cognitive_Priority"] == "High") & (df["SP_Empathy_Threshold"] == "High")],
            lambda df: df[(df["SP_Motivational_Priority"].isin(["High", "Medium"])) & (df["SP_Empathy_Threshold"].isin(["Low", "Medium"]))],
            lambda df: df[(df["SP_Motivational_Priority"].isin(["High", "Medium"])) & (df["SP_Empathy_Threshold"] == "High")],
        ],
    },
]


def plot_radar_grid(
    models: Dict[str, Path],
    output_dir: str | Path = ".",
    *,
    chart_ids: Optional[Sequence[str]] = None,
) -> List[Path]:
    """Generate radar-grid small-multiples for model comparison.

    Parameters
    ----------
    models : dict
        ``{display_name: results_dir}``.
    output_dir : path-like
        Where to save outputs.
    chart_ids : list of str, optional
        Which radar configs to generate. Default: all three
        (``"categories"``, ``"mechanism"``, ``"persona"``).

    Returns
    -------
    list of Path
        Saved PNG paths.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    _apply_academic_style(plt)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    model_names = list(models.keys())
    colors = _assign_colors(model_names)
    n_models = len(model_names)

    configs = _RADAR_CONFIGS
    if chart_ids:
        configs = [c for c in configs if c["id"] in chart_ids]

    saved: List[Path] = []

    for config in configs:
        labels = config["labels"]
        N = len(labels)
        angles = [n / float(N) * 2 * pi for n in range(N)] + [0]

        # Compute scores per model
        data: Dict[str, List[float]] = {}
        for name, path in models.items():
            scores = _radar_scores(path, config["filters"])
            if scores:
                data[name] = scores

        if not data:
            print(f"⚠️  No data for {config['id']}, skipping")
            continue

        all_scores = [s for scores in data.values() for s in scores]
        limit = ceil(max(all_scores) / 20) * 20 + 20
        if limit < 100:
            limit = 100

        cols = min(4, n_models)
        rows = ceil(n_models / cols)
        fig, axes_arr = plt.subplots(rows, cols, subplot_kw=dict(polar=True),
                                     figsize=(5.5 * cols, 6 * rows))
        if n_models == 1:
            axes_flat = [axes_arr]
        else:
            axes_flat = list(np.array(axes_arr).flatten())

        fig.suptitle(config["title"], fontsize=22, fontweight="bold", y=0.995,
                     color="#9E2236", fontfamily="Times New Roman")
        if config.get("subtitle"):
            fig.text(0.5, 0.97, config["subtitle"], ha="center", va="top",
                     fontsize=12, color="#444", style="italic",
                     fontfamily="Times New Roman", linespacing=1.4)

        ordered_names = [m for m in model_names if m in data]

        for i, ax in enumerate(axes_flat):
            if i >= len(ordered_names):
                ax.axis("off")
                continue

            mname = ordered_names[i]
            scores = data[mname]
            values = scores + scores[:1]
            color = colors.get(mname, "grey")

            ax.set_theta_offset(pi / 2)
            ax.set_theta_direction(-1)
            ax.set_xticks([n / float(N) * 2 * pi for n in range(N)])
            ax.set_xticklabels(labels, fontsize=10, color="#9E2236", fontweight="bold",
                               fontfamily="Times New Roman")
            plt.setp(ax.get_yticklabels(), visible=False)
            ax.set_ylim(0, limit)
            ax.spines["polar"].set_visible(False)

            ax.plot(angles, values, lw=2.2, color=color)
            ax.fill(angles, values, color=color, alpha=0.28)
            ax.text(0.5, -0.10, f"[{i + 1}] {mname}", transform=ax.transAxes,
                    ha="center", va="top", fontsize=11, fontweight="bold",
                    fontfamily="Times New Roman")
            avg = np.mean(scores)
            ax.text(0, 0, f"{avg:.1f}", ha="center", va="center",
                    fontsize=12, fontweight="bold", color="#9E2236")

        plt.tight_layout(rect=[0, 0.02, 1, 0.94])
        out = output_dir / config["filename"]
        plt.savefig(str(out), dpi=300, bbox_inches="tight")
        plt.savefig(str(out).replace(".png", ".pdf"), bbox_inches="tight")
        plt.close(fig)
        print(f"✅ Radar grid ({config['id']}): {out}")
        saved.append(out)

    return saved


def _radar_scores(model_dir: Path, filters: list) -> Optional[List[float]]:
    csv = model_dir / "descriptive_statistics.csv"
    if not csv.exists():
        return None
    try:
        df = _normalize_df(pd.read_csv(csv))
        df = df[df["Case_ID"].astype(str).str.startswith("script_", na=False)]

        _enrich_sp_metadata(df)

        df["EPM_Index"] = df.apply(
            lambda r: 0.4 * ((r.get("Idx_RDI", 0) + r.get("Idx_Etot", 0) + r.get("Idx_Snet", 0)) / 3)
            + 0.2 * ((r.get("Idx_Rho", 0) + r.get("Idx_Sproj", 0) + r.get("Idx_Tau", 0)) / 3)
            + 0.4 * ((r.get("Idx_Rpos", 0) + r.get("Idx_Align", 0) + r.get("Idx_Pen", 0)) / 3),
            axis=1,
        )
        scores = []
        for fn in filters:
            try:
                sub = fn(df)
                scores.append(sub["EPM_Index"].mean() if len(sub) > 0 else 0)
            except Exception:
                scores.append(0)
        return [0 if np.isnan(x) else x for x in scores]
    except Exception:
        return None


def _enrich_sp_metadata(df: pd.DataFrame) -> None:
    """Fill missing SP metadata columns by extracting from actor prompt files."""
    sp_cols = ["SP_Empathy_Threshold", "SP_Emotional_Priority",
               "SP_Motivational_Priority", "SP_Cognitive_Priority"]
    needs_fill = any(
        col not in df.columns or df[col].isna().all() for col in sp_cols
    )
    if not needs_fill:
        return

    from empa.evaluation.epmq import extract_sp_metadata

    key_col = "Case_ID" if "Case_ID" in df.columns else "Case_ID"
    mapping = {
        "sp_threshold_level": "SP_Empathy_Threshold",
        "sp_emotional_priority": "SP_Emotional_Priority",
        "sp_motivational_priority": "SP_Motivational_Priority",
        "sp_cognitive_priority": "SP_Cognitive_Priority",
    }

    for dst_col in sp_cols:
        if dst_col not in df.columns:
            df[dst_col] = pd.Series(dtype="object")
        else:
            df[dst_col] = df[dst_col].astype("object")

    for idx, row in df.iterrows():
        sid = str(row[key_col])
        sp = extract_sp_metadata(sid)
        for src_key, dst_col in mapping.items():
            val = sp.get(src_key, "N/A")
            df.at[idx, dst_col] = _VAL_SP.get(val, val)


def _translate_reason(reason: str) -> str:
    """Translate a Chinese termination reason to English."""
    import re as _re
    if not reason:
        return ""
    # Extract the part after ":"
    if ":" in reason:
        reason = reason.split(":")[-1].strip()

    _MAP = [
        ("位置胜利（成功穿越或接近目标区域）", "Positional Victory"),
        ("几何胜利（精准到达原点附近）", "Geometric Victory"),
        ("连续5轮负能量（方向崩溃）", "Direction Collapse (5 consecutive neg-energy turns)"),
        ("持续倒退（8轮中70%负能量）", "Persistent Regression (70% neg in 8 turns)"),
        ("对话停滞", "Dialogue Stagnation"),
    ]
    result = reason
    for cn, en in _MAP:
        result = result.replace(cn, en)

    m = _re.search(r"向量反复震荡[：:]\s*总变化(\d+).*净变化仅(\d+)", result)
    if m:
        result = f"Vector Oscillation (total change {m.group(1)}, net change {m.group(2)})"

    m = _re.search(r"达到最大轮次\((\d+)\).*对话超时", result)
    if m:
        result = f"Max turns ({m.group(1)}) reached, timeout"

    return result


# ---------------------------------------------------------------------------
# Cross-model comparison table
# ---------------------------------------------------------------------------

def generate_comparison_table(
    models: Dict[str, Path],
    output_dir: str | Path = ".",
) -> Path:
    """Generate a cross-model per-case comparison table (CSV + Excel).

    Produces a table with case metadata on the left and per-model
    Result / Reason / Turns columns, similar to the paper's
    ``MODEL_COMPARISON_BY_CASE.xlsx``.

    Parameters
    ----------
    models : dict
        ``{display_name: results_dir}``.
    output_dir : path-like
        Where to save outputs.

    Returns
    -------
    Path
        Saved Excel path.
    """
    import json as _json

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    from empa.evaluation.report import load_case_metadata
    case_meta = load_case_metadata()

    case_ids: list[str] = []
    first_model_dir = next(iter(models.values()))
    for jf in sorted(first_model_dir.glob("script_*_result.json")):
        sid = jf.stem.replace("_result", "")
        case_ids.append(sid)
    if not case_ids:
        for jf in sorted(first_model_dir.glob("script_*.json")):
            sid = jf.stem
            if sid.startswith("script_"):
                case_ids.append(sid)

    _DIFF_MAP = {"较易": "Easy", "中等": "Medium", "困难": "Hard", "极难": "Very Hard"}
    _CAT_MAP = {
        "休闲娱乐": "Leisure", "身心健康": "Health", "生活状况": "Life",
        "观念认同": "Values", "人际关系": "Relations", "生涯发展": "Career",
    }

    rows: list[dict] = []
    for sid in case_ids:
        meta = case_meta.get(sid, {})
        row: dict[str, Any] = {
            ("Case Metadata", "Case ID"): sid,
            ("Case Metadata", "Difficulty"): _DIFF_MAP.get(meta.get("difficulty", ""), meta.get("difficulty", "")),
            ("Case Metadata", "Category"): _CAT_MAP.get(meta.get("category", ""), meta.get("category", "")),
        }

        for model_name, model_dir in models.items():
            result_file = model_dir / f"{sid}_result.json"
            if not result_file.exists():
                result_file = model_dir / f"{sid}.json"

            if result_file.exists():
                with open(result_file, "r", encoding="utf-8") as f:
                    data = _json.load(f)
                turns = data.get("total_turns", 0)
                reason = data.get("termination_reason", "")

                if "判定成功" in reason:
                    result_str = "✅ Success"
                elif "超时" in reason or "最大轮次" in reason:
                    result_str = "⚠️ Timeout"
                else:
                    result_str = "❌ Failure"

                reason_short = _translate_reason(reason)
            else:
                result_str = "N/A"
                reason_short = ""
                turns = ""

            row[(model_name, "Result")] = result_str
            row[(model_name, "Reason")] = reason_short
            row[(model_name, "Turns")] = turns

        rows.append(row)

    df = pd.DataFrame(rows)
    df.columns = pd.MultiIndex.from_tuples(df.columns)

    csv_path = output_dir / "model_comparison_by_case.csv"
    df_flat = df.copy()
    df_flat.columns = [f"{a} | {b}" if a != b else a for a, b in df.columns]
    df_flat.to_csv(str(csv_path), index=False, encoding="utf-8-sig")
    print(f"✅ Comparison table CSV: {csv_path}")

    xlsx_path = output_dir / "model_comparison_by_case.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        center_align = Alignment(horizontal="center", wrap_text=True)

        level0 = [c[0] for c in df.columns]
        level1 = [c[1] for c in df.columns]

        for ci, val in enumerate(level0, 1):
            cell = ws.cell(row=1, column=ci, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for ci, val in enumerate(level1, 1):
            cell = ws.cell(row=2, column=ci, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        i = 0
        while i < len(level0):
            j = i + 1
            while j < len(level0) and level0[j] == level0[i]:
                j += 1
            if j - i > 1:
                ws.merge_cells(
                    start_row=1, start_column=i + 1,
                    end_row=1, end_column=j,
                )
            i = j

        for ri, (_, row_data) in enumerate(df.iterrows(), 3):
            for ci, val in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=val)

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 22)

        wb.save(str(xlsx_path))
        print(f"✅ Comparison table Excel: {xlsx_path}")
    except ImportError:
        print("⚠️  openpyxl not installed, skipping Excel. Install with: pip install openpyxl")

    return xlsx_path


# ---------------------------------------------------------------------------
# EPM-Q summary table (merged from 3 separate tables)
# ---------------------------------------------------------------------------

_IDX_LABELS = {
    "Idx_RDI": "RDI",
    "Idx_Etot": "E_total",
    "Idx_Snet": "S_net",
    "Idx_Rho": "ρ",
    "Idx_Sproj": "S_proj",
    "Idx_Tau": "τ",
    "Idx_Rpos": "R_pos",
    "Idx_Align": "Align",
    "Idx_Pen": "R_pen",
}

_RAW_COL_FOR_IDX = {
    "Idx_RDI": "Distance_Improvement%",
    "Idx_Etot": "Total_Energy",
    "Idx_Snet": "Total_Net_Score",
    "Idx_Rho": "Net_Score_Per_Turn",
    "Idx_Sproj": "Avg_Effective_Projection",
    "Idx_Tau": "Tortuosity",
    "Idx_Rpos": "Positive_Energy_Ratio%",
    "Idx_Align": "Avg_Alignment",
    "Idx_Pen": "Penalty_Rate",
}

_GROUP_FOR_IDX = {
    "Idx_RDI": "Outcome Quality", "Idx_Etot": "Outcome Quality", "Idx_Snet": "Outcome Quality",
    "Idx_Rho": "Process Efficiency", "Idx_Sproj": "Process Efficiency", "Idx_Tau": "Process Efficiency",
    "Idx_Rpos": "Process Stability", "Idx_Align": "Process Stability", "Idx_Pen": "Process Stability",
}


def generate_epmq_summary_table(
    models: Dict[str, Path],
    output_dir: str | Path = ".",
) -> Path:
    """Generate a single comprehensive EPM-Q summary table.

    Merges three previously separate tables into one:
    - Standardized index scores (9 indices)
    - Category composites (Outcome, Efficiency, Stability)
    - Final EPM-Q Score

    The resulting Excel file has two sheets:
    - **Standardized Scores**: 9 indices + 3 composites + EPM-Q Score
    - **Raw Statistics**: mean ± std for the 9 underlying metrics

    Models are sorted by EPM-Q Score descending.

    Returns
    -------
    Path
        Saved Excel path.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    long_df = _load_all(models)
    model_names = list(models.keys())

    idx_order = list(IDX_COLUMNS)
    group_names = ["Outcome Quality", "Process Efficiency", "Process Stability"]

    rows_std: list[dict] = []
    rows_raw: list[dict] = []

    for mname in model_names:
        mdf = long_df[long_df["Model"] == mname]
        if mdf.empty:
            continue

        row_s: dict = {"Model": mname}
        row_r: dict = {"Model": mname}
        group_scores: dict[str, list[float]] = {g: [] for g in group_names}

        for idx_col in idx_order:
            if idx_col in mdf.columns:
                vals = mdf[idx_col].astype(float)
                score = vals.mean()
            else:
                score = 0.0
            row_s[idx_col] = round(score, 2)
            group_scores[_GROUP_FOR_IDX[idx_col]].append(score)

            raw_col = _RAW_COL_FOR_IDX.get(idx_col, "")
            if raw_col and raw_col in mdf.columns:
                rvals = mdf[raw_col].astype(float)
                row_r[idx_col] = f"{rvals.mean():.2f} ± {rvals.std():.2f}"
            else:
                row_r[idx_col] = "N/A"

        composites = {}
        for g in group_names:
            composites[g] = np.mean(group_scores[g]) if group_scores[g] else 0.0
            row_s[g] = round(composites[g], 2)

        epm_q = (
            W_OUTCOME * composites["Outcome Quality"]
            + W_EFFICIENCY * composites["Process Efficiency"]
            + W_STABILITY * composites["Process Stability"]
        )
        row_s["EPM-Q Score"] = round(epm_q, 2)

        rows_std.append(row_s)
        rows_raw.append(row_r)

    df_std = pd.DataFrame(rows_std).sort_values("EPM-Q Score", ascending=False).reset_index(drop=True)
    df_raw = pd.DataFrame(rows_raw)
    df_raw = df_raw.set_index("Model").loc[df_std["Model"]].reset_index()

    csv_path = output_dir / "epmq_summary.csv"
    df_std.to_csv(str(csv_path), index=False, encoding="utf-8-sig")
    print(f"✅ EPM-Q summary CSV: {csv_path}")

    xlsx_path = output_dir / "epmq_summary.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Standardized Scores ──
        ws1 = wb.active
        ws1.title = "Standardized Scores"

        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            bottom=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="D0D0D0"),
        )

        col_spec: list[tuple[str, str, str]] = [("", "Model", "Model")]
        for idx_col in idx_order:
            col_spec.append((_GROUP_FOR_IDX[idx_col], _IDX_LABELS[idx_col], idx_col))
        for g in group_names:
            col_spec.append(("Composites", g, g))
        col_spec.append(("Composites", "EPM-Q Score", "EPM-Q Score"))

        # Row 1: group headers
        for ci, (grp, _, _) in enumerate(col_spec, 1):
            cell = ws1.cell(row=1, column=ci, value=grp)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
        # Merge group headers
        i = 0
        while i < len(col_spec):
            grp = col_spec[i][0]
            j = i + 1
            while j < len(col_spec) and col_spec[j][0] == grp:
                j += 1
            if j - i > 1 and grp:
                ws1.merge_cells(start_row=1, start_column=i + 1, end_row=1, end_column=j)
            i = j

        # Row 2: column headers
        for ci, (_, label, _) in enumerate(col_spec, 1):
            cell = ws1.cell(row=2, column=ci, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center

        # Data rows
        for ri, (_, row_data) in enumerate(df_std.iterrows(), 3):
            for ci, (_, _, key) in enumerate(col_spec, 1):
                val = row_data.get(key, "")
                cell = ws1.cell(row=ri, column=ci, value=val)
                cell.alignment = center
                cell.border = thin_border

        # Column widths
        for ci in range(1, len(col_spec) + 1):
            letter = get_column_letter(ci)
            max_len = max(
                len(str(ws1.cell(row=r, column=ci).value or ""))
                for r in range(1, ws1.max_row + 1)
            )
            ws1.column_dimensions[letter].width = min(max(max_len + 2, 10), 22)

        # ── Sheet 2: Raw Statistics ──
        ws2 = wb.create_sheet("Raw Statistics")

        _RAW_LABELS_CLEAN = {
            "Idx_RDI": "RDI (%)", "Idx_Etot": "E_total", "Idx_Snet": "S_net",
            "Idx_Rho": "ρ (energy/turn)", "Idx_Sproj": "S_proj", "Idx_Tau": "τ (tortuosity)",
            "Idx_Rpos": "R_pos (%)", "Idx_Align": "Avg cos(θ)", "Idx_Pen": "R_pen",
        }
        raw_col_spec: list[tuple[str, str, str]] = [("", "Model", "Model")]
        for idx_col in idx_order:
            raw_label = _RAW_LABELS_CLEAN.get(idx_col, _IDX_LABELS[idx_col])
            raw_col_spec.append((_GROUP_FOR_IDX[idx_col], raw_label, idx_col))

        for ci, (grp, _, _) in enumerate(raw_col_spec, 1):
            cell = ws2.cell(row=1, column=ci, value=grp)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
        i = 0
        while i < len(raw_col_spec):
            grp = raw_col_spec[i][0]
            j = i + 1
            while j < len(raw_col_spec) and raw_col_spec[j][0] == grp:
                j += 1
            if j - i > 1 and grp:
                ws2.merge_cells(start_row=1, start_column=i + 1, end_row=1, end_column=j)
            i = j

        for ci, (_, label, _) in enumerate(raw_col_spec, 1):
            cell = ws2.cell(row=2, column=ci, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center

        for ri, (_, row_data) in enumerate(df_raw.iterrows(), 3):
            for ci, (_, _, key) in enumerate(raw_col_spec, 1):
                val = row_data.get(key, "")
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.alignment = center
                cell.border = thin_border

        for ci in range(1, len(raw_col_spec) + 1):
            letter = get_column_letter(ci)
            max_len = max(
                len(str(ws2.cell(row=r, column=ci).value or ""))
                for r in range(1, ws2.max_row + 1)
            )
            ws2.column_dimensions[letter].width = min(max(max_len + 2, 10), 22)

        wb.save(str(xlsx_path))
        print(f"✅ EPM-Q summary Excel: {xlsx_path}")
    except ImportError:
        print("⚠️  openpyxl not installed, skipping Excel.")

    return xlsx_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _short_name(name: str, max_len: int = 14) -> str:
    if len(name) <= max_len:
        return name
    parts = name.split()
    if len(parts) >= 2:
        return parts[0][:6] + " " + parts[-1]
    return name[:max_len]


def _apply_academic_style(plt_module: Any) -> None:
    plt_module.rcParams.update({
        "font.family": "serif",
        "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
        "mathtext.fontset": "stix",
        "font.size": 10,
        "axes.labelsize": 10,
        "axes.titlesize": 12,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "legend.fontsize": 9,
        "axes.linewidth": 0.75,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "savefig.dpi": 300,
        "savefig.bbox": "tight",
        "legend.frameon": False,
    })
